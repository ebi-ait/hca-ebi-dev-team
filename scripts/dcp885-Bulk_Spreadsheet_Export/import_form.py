import datetime

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

R24_PROJECTS = [
    'cc95ff89-2e68-4a08-a234-480eca21ce79',
    '258c5e15-d125-4f2d-8b4c-e3122548ec9b',
    'cdc2d270-6c99-4142-8883-9bd95c041d05',
    '77dedd59-1376-4887-9bca-dc42b56d5b7a',
    'e88714c2-2e78-49da-8146-5a60b50628b4',
    'e57dc176-ab98-446b-90c2-89e0842152fd',
    '8787c238-89ef-4636-a57d-3167e8b54a80',
    '5f607e50-ba22-4598-b1e9-f3d9d7a35dcc',
    'cc95ff89-2e68-4a08-a234-480eca21ce79',
    '957261f7-2bd6-4358-a6ed-24ee080d5cfc',
    '3ce9ae94-c469-419a-9637-5d138a4e642f',
    'cbd2911f-252b-4428-abde-69e270aefdfc',
    '2043c65a-1cf8-4828-a656-9e247d4e64f1',
    'cd9d6360-ce38-4321-97df-f13c79e3cb84',
    '95d058bc-9cec-4c88-8d2c-05b4a45bf24f',
    'cd9d6360-ce38-4321-97df-f13c79e3cb84',
    'cdabcf0b-7602-4abf-9afb-3b410e545703',
    'f2078d5f-2e7d-4844-8552-f7c41a231e52',
    '2d4d89f2-ebeb-467c-ae60-a3efc5e8d4ba',
    '12f32054-8f18-4dae-8959-bfce7e3108e7',
    '2d4d89f2-ebeb-467c-ae60-a3efc5e8d4ba',
    '6e60a555-fd95-4aa2-8e29-3ec2ef01a580'
]


class ImportForm:
    def __init__(self):
        self.workbook = Workbook()
        self.success = self.workbook.create_sheet("Import Form")
        self.errors = self.workbook.create_sheet("Errors")
        self.__init_import_form(self.success)

    @staticmethod
    def __init_import_form(sheet: Worksheet):
        headers = [
            'Timestamp', 'Email Address', 'Institution', 'Project UUID', 'Environment', 'Catalog', 'Dataset ID',
            'Create a snapshot?', 'Additional Info', 'Release #', 'Is this new or updated data? (check all that apply)',
            'Project Name', 'Optional Release Note for HCA Releases', 'Your Institution', 'Verified by Wrangler'
        ]
        sheet.append(headers)

    def add_success(self, project_uuid: str, project_name, release='25'):
        timestamp = datetime.datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')
        email = 'wrangler-team@data.humancellatlas.org'
        institute = 'EBI'
        environment = 'prod'
        catalogue = ''
        dataset = ''
        snapshot = ''
        additional = 'Metadata Spreadsheet'
        new_or_updated = 'Updated metadata spreadsheet'
        release_notes = ''
        verified = ''
        row = [
            timestamp, email, institute, project_uuid, environment, catalogue, dataset,
            snapshot, additional, release, new_or_updated,
            project_name, release_notes, institute, verified
        ]
        self.success.append(row)
