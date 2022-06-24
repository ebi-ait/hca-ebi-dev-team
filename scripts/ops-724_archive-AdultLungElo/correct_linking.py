import json
import os

import requests
from ingest.api.ingestapi import IngestApi
from openpyxl import load_workbook

# download xlsx from https://docs.google.com/spreadsheets/u/2/d/1-mNz7hkdVovp8QDgeTIZK4Yp8q3bu0101URsrSWDptY/edit#gid=0
workbook = load_workbook(filename=f'{os.getcwd()}/_local/AdultLungElo-Linking.xlsx')
sheet = workbook.active

TOKEN = 'insert-bearer-token-with-prefix'
INGEST_API = 'https://api.ingest.archive.data.humancellatlas.org'
ingest_api = IngestApi(url=INGEST_API)
ingest_api.set_token(f'Bearer {TOKEN}')
headers = {
    'Content-type': 'text/uri-list',
    'Authorization': f'Bearer {TOKEN}'
}


def link_to_entity(from_entity, process, to_entity, protocols):
    ingest_api.link_entity(from_entity, process, 'inputToProcesses')
    ingest_api.link_entity(to_entity, process, 'derivedByProcesses')

    for protocol in protocols:
        ingest_api.link_entity(process, protocol, 'protocols')


def unlink_entities(from_entity, relationship, to_entity):
    from_uri = from_entity["_links"][relationship]["href"]
    to_uri = to_entity["_links"]['self']["href"]
    to_id = to_uri.split('/')[-1]

    delete_url = f'{from_uri}/{to_id}'
    # using ingest_api.session so that it retries
    r = ingest_api.session.delete(delete_url, headers=headers)
    r.raise_for_status()
    print(f'removed link {delete_url}')


def delete_entity(entity):
    uri = entity["_links"]['self']["href"]
    r = requests.delete(uri, headers=headers)
    print(f'removed entity {uri}')
    r.raise_for_status()


def write_json(obj, path, **kwargs):
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(obj, f, **kwargs)


data = {}
for row in sheet.iter_rows(min_row=2, values_only=True):
    sequence_file_name = row[1]
    sequence_file_data = {
        'sequence_file.uuid': row[0],
        'process.uuid': row[2],
        'cell_suspension.uuid': row[5],
        'library_preparation_protocol.uuid': row[6],
        'sequencing_protocol.uuid': row[7]
    }
    data[sequence_file_name] = sequence_file_data

process_uuids_to_unlink = {}
grouping = {}
for sequence_file_name, sequence_file_data in data.items():
    prefix = sequence_file_name.split('_S1_')[0]
    group = grouping.get(prefix, {})

    process_uuid = sequence_file_data['process.uuid']
    sequence_file_uuid = sequence_file_data['sequence_file.uuid']
    cell_susp_uuid = sequence_file_data['cell_suspension.uuid']
    libprep_protocol_uuid = sequence_file_data['library_preparation_protocol.uuid']
    seq_protocol_uuid = sequence_file_data['sequencing_protocol.uuid']

    # use the first process.uuid of the group as assay process
    if not group.get('process.uuid'):
        group['process.uuid'] = process_uuid
    else:
        process_uuids_to_unlink[process_uuid] = {}
        process_uuids_to_unlink[process_uuid]['process.uuid'] = process_uuid
        process_uuids_to_unlink[process_uuid]['sequence_file.uuid'] = sequence_file_uuid
        process_uuids_to_unlink[process_uuid]['cell_suspension.uuid'] = cell_susp_uuid
        process_uuids_to_unlink[process_uuid]['library_preparation_protocol.uuid'] = libprep_protocol_uuid
        process_uuids_to_unlink[process_uuid]['sequencing_protocol.uuid'] = seq_protocol_uuid

        file_uuids = group.get('files', [])
        file_uuids.append(sequence_file_uuid)
        group['files'] = file_uuids

    curr_cell_susp_in_grp = group.get('cell_suspension.uuid')
    if curr_cell_susp_in_grp and curr_cell_susp_in_grp != cell_susp_uuid:
        message = f'file {sequence_file_uuid} cell suspension uuid {cell_susp_uuid} but has same prefix {prefix}'
        raise ValueError(message)
    else:
        group['cell_suspension.uuid'] = cell_susp_uuid

    grouping[prefix] = group

files_by_process_uuid = {}
for prefix, group in grouping.items():
    files_by_process_uuid[group['process.uuid']] = group

write_json(files_by_process_uuid, 'files_by_process_uuid.json')
write_json(process_uuids_to_unlink, 'process_uuids_to_unlink.json')

for process_uuid, linked_data in process_uuids_to_unlink.items():
    try:
        process = ingest_api.get_entity_by_uuid('processes', process_uuid)
    except Exception as e:
        print(f'{process_uuid} already deleted')
        continue

    cell_susp_uuid = linked_data['cell_suspension.uuid']
    cell_suspension = ingest_api.get_entity_by_uuid('biomaterials', cell_susp_uuid)
    unlink_entities(cell_suspension, 'inputToProcesses', process)

    file_uuid = linked_data['sequence_file.uuid']
    file = ingest_api.get_entity_by_uuid('files', file_uuid)
    unlink_entities(file, 'derivedByProcesses', process)

    libprep_protocol_uuid = linked_data['library_preparation_protocol.uuid']
    libprep_protocol = ingest_api.get_entity_by_uuid('protocols', libprep_protocol_uuid)
    unlink_entities(process, 'protocols', libprep_protocol)

    seq_protocol_uuid = linked_data['sequencing_protocol.uuid']
    seq_protocol = ingest_api.get_entity_by_uuid('protocols', seq_protocol_uuid)
    unlink_entities(process, 'protocols', seq_protocol)
    try:
        delete_entity(process)
    except Exception as e:
        print(f'{process_uuid} already deleted')

for process_uuid, linked_data in files_by_process_uuid.items():
    process = ingest_api.get_entity_by_uuid('processes', process_uuid)
    files = linked_data['files']
    for file_uuid in files:
        file = ingest_api.get_entity_by_uuid('files', file_uuid)
        ingest_api.link_entity(file, process, 'derivedByProcesses')
        print(f'link process {process_uuid} to {file_uuid}')
