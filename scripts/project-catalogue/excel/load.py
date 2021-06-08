from contextlib import closing
from typing import Tuple
from openpyxl import load_workbook

from excel_submission_broker.submission import ExcelSubmission, Entity


class SheetLoader:
    def __init__(self, excel_path: str, sheet=None):
        self.sheet = sheet if sheet else 0
        with closing(load_workbook(filename=excel_path, read_only=True, keep_links=False)) as workbook:
            worksheet = workbook[self.sheet]
            self.column_map, self.data = self.get_sheet(worksheet)

    def get_sheet(self, worksheet) -> Tuple[dict, ExcelSubmission]:
        data = ExcelSubmission()
        header_row = True
        column_map = {}
        row_index = 1
        # Import cell values into data object
        # Uses .iter_rows for faster reads, requires workbook read_only=True
        for row in worksheet.iter_rows():
            if header_row:
                self.map_header_row(column_map, row)
                header_row = False
            else:
                json_row = self.map_data_row(column_map, row)
                self.map_row_entity(data, row_index, json_row)
            row_index += 1
        return column_map, data

    def map_row_entity(self, submission: ExcelSubmission, row_index: int, attributes: dict):
        index = self.get_index(attributes, row_index)
        return submission.map_row(row_index, self.sheet, index, attributes)
    
    @staticmethod
    def map_header_row(column_map: dict, row):
        for header in row:
            if isinstance(header.value, str) and header.value.strip():
                column_map[header.column_letter] = header.value.strip().lower()

    @staticmethod
    def map_data_row(column_map: dict, row):
        row_data = {}
        for cell in row:
            if cell.value is not None:
                if cell.is_date:
                    value = cell.value.date().isoformat()
                else:
                    value = str(cell.value).strip()

                if cell.column_letter in column_map:
                    attribute_name = column_map[cell.column_letter]
                    row_data[attribute_name] = value
        return row_data

    @staticmethod
    def get_index(attributes: dict, row: int) -> str:
        if 'dcp_id' in attributes:
            return attributes['dcp_id']
        else:
            return f'row_{row}'
